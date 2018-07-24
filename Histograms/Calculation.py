import csv
import numpy
import matplotlib
import matplotlib.pyplot as plt
import openpyxl as xlsx
import os

def is_number(s):
    try:
        float(s)
        return True
    except:
        return False

stimuliCount = 50;
stimuliPeriods = [5, 10, 50, 500];

#dataFolder = r"C:\Projects\Misa\Data\eIPSC\tt";
#dataFolder = r"C:\Projects\Misa\Data\eIPSC\vyhodnoceni Clampfit";
dataFolder = r"C:\Projects\Misa\vyhodnoceni EGTA-1st";
outputFolder = r"C:\Projects\Misa\vyhodnoceni EGTA-1st\Out";

if not os.path.exists(outputFolder):
    os.mkdir(outputFolder);

for filepath in os.listdir(dataFolder):
#filepathRoot = r"C:\Projects\Misa\Data\eIPSC\vyhodnoceni Gtest\a161116_";

#for i in range(1, 80):
 #   filepath = filepathRoot + '{:04}'.format(i) + ".xlsx";

    if not filepath.endswith(".xlsx"):
        continue;

    if filepath.startswith('~'):
        continue;

    print(filepath);

    filepath = os.path.join(dataFolder, filepath);
    workbook = xlsx.load_workbook(filepath);

    filename = os.path.basename(filepath);
    filenameWithoutExt = os.path.splitext(filename)[0];

    #ws = workbook.get_sheet_by_name(filenameWithoutExt);
#    if len(workbook.worksheets) != 1 :
 #       raise Exception('Too many worksheets!');

    ws = None;
    for wsItem in workbook.worksheets:
        if str(wsItem.title.upper()).startswith(filename.upper()[0]):
            ws = wsItem;
            break;

    if ws == None:
        raise("Cannot find a valid worksheet - its name should be the same as the filename")


    uniqueStimuli = set();

    nominalFrequency = float(ws["A1"].value);

    for col in ws.iter_cols(min_col=2 , max_col=2, min_row=1, max_row=10000):
        for cell in col:
            if is_number(cell.value):
                uniqueStimuli.add(float(cell.value));

    uniqueStimuli = sorted(uniqueStimuli);

#    if (len(uniqueStimuli) != stimuliCount):
 #       raise Exception('Stimuli count is {0}. Expected {1}!'.format(len(uniqueStimuli), stimuliCount));

    detectedIntervals = numpy.diff(uniqueStimuli);

    counts = [0] * len(stimuliPeriods);
    counts[0] = 1;
    for detectedInterval in detectedIntervals:
        idx = (numpy.abs(stimuliPeriods-detectedInterval)).argmin();
        counts[idx] = counts[idx] + 1;

    aa = numpy.abs(counts).argmax();
    detectedStimuliPeriod = stimuliPeriods[aa];

    if 1/(detectedStimuliPeriod * 1e-3) !=  nominalFrequency:
        raise "Different nominal and detected frequency";

    #for stimuliPeriod in stimuliPeriods:
    #histogramStimuli, bin_edges_stimuli = numpy.histogram(detectedIntervals, stimuliBins);

#detectedStimuliPeriod = uniqueStimuli[1]-uniqueStimuli[0];

#stimuliPeriodFound = False;

    #for stimuliPeriod in stimuliPeriods:
#    if abs(stimuliPeriod-detectedStimuliPeriod) < stimuliPeriod*0.1:
#           detectedStimuliPeriod = stimuliPeriod;
#           stimuliPeriodFound = True;
#           break;

#   if (detectedStimuliPeriodTmp != detectedStimuliPeriod):
#       raise Exception("Stimulation period different!")

#   if not stimuliPeriodFound:
#       raise Exception("Stimulation period not detected!")


    #generate bins for the histogram
    stimuli = [uniqueStimuli[0] + x * detectedStimuliPeriod for x in range(0, stimuliCount+1)]

    #stimuli = numpy.arange(uniqueStimuli[0], uniqueStimuli[0]+detectedStimuliPeriod*(stimuliCount+1), detectedStimuliPeriod);
    #print(stimuli)

#    intervals = events[1:len(events)-1] - events[0:len(events)-2];

    events = list();
    for col in ws.iter_cols(min_col=3 , max_col=3, min_row=1, max_row=10000):
        for cell in col:
            if (is_number(cell.value)):
                events.append(float(cell.value));

    events = sorted(events);

    histogram, bin_edges = numpy.histogram(events, stimuli);

    cumulativeHistogram = numpy.cumsum(histogram);
    splitEvents = numpy.split(events, cumulativeHistogram);

    for item in splitEvents:
        if len(item) > 1:
            intervals = numpy.mean(numpy.diff(item));

    binIndices = numpy.digitize(events, bin_edges)


    #x=[x/1000 for x in random.sample(range(100000),100)]
    #xbins=range(0,len(x))
    #plt.hist(x, bins=xbins, color='blue')

    left = bin_edges[0:len(bin_edges)-1];
    widths = (bin_edges[1:len(bin_edges)] - left);

    #dirName = os.path.dirname(filepath);
    dirName = outputFolder;

    saveGraphFigure = False;
    if saveGraphFigure:
        fig = plt.bar(left, histogram, width=widths);

        plt.savefig(os.path.join(dirName, "hist_" + filenameWithoutExt + "_{0}Hz.png".format((int)(1000/detectedStimuliPeriod))));

        plt.cla();

    #save the histogram
    file = open(
        os.path.join(dirName, "hist_" + filenameWithoutExt + "_{0}Hz.txt".format((int)(1000 / detectedStimuliPeriod))),
        mode='w');

    for i in range(0, len(histogram)):
        file.write("{0}\t{1}\n".format(bin_edges[i], histogram[i]));

    file.write("{0}".format(bin_edges[len(bin_edges) - 1]));

    file.close();

    continue;

    if histogram[0] != 0:
        normalizationConstant = float(histogram[0]);
    else:
        normalizationConstant = float(histogram[1]);

    histogram = histogram / normalizationConstant;

    fig = plt.bar(left, histogram, width=widths);

    plt.savefig(os.path.join(dirName, "hist_norm_" + filenameWithoutExt + ".png"));

    file = open(os.path.join(dirName, "hist_norm_" + filenameWithoutExt + ".txt"), mode='w');

    for i in range(0, len(histogram)):
        file.write("{0}\t{1}\n".format(bin_edges[i], histogram[i]));

    file.write("{0}".format(bin_edges[len(bin_edges)-1]));

    file.close();
    plt.cla();

    #plt.hist(bin_edges[0:5], weights=aa);
    #plt.hist(aa[0], bins=intervals, color='blue');
    #plt.show()

    #plt.cla();
    #plt.clf()

    #plt.hist([1, 2, 3, 4, 5, 6],weights=aa[0])


    #plt.bar(aa[0], bin_edges);
    #aa = plt.hist(data, intervals);
    #plt.bar([1, 2, 3, 4, 5], aa);
    #plt.show();


"""with open("C:\Projects\Misa\Cell1.txt") as dataFile:
    readerData = csv.reader(dataFile, delimiter='\n');

    with open("C:\Projects\Misa\intervals.txt") as intervalsFile:
        readerIntervals = csv.reader(intervalsFile);

        data = list();

        for row in readerData:
            for number in row:
                data.append(float(number));

        intervals = list();
        for row in readerIntervals:
            for itemData in row:
                intervals.append(float(itemData));
                # </editor-fold>
                # endregion
"""